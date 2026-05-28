# -*- coding: utf-8 -*-
"""
Exportación a Excel — SIATC
Ministerio Público Fiscal de Córdoba
"""

import pandas as pd
from io import BytesIO
from datetime import date, datetime
import database as db
from data_cordoba import TIPOS_INFRACCION, UNIDADES

CARRIL_LABEL = {"verde": "Mediación", "amarillo": "Suspensión", "rojo": "Proceso pleno"}
TIPO_RES_LABEL = {"suspension": "Suspensión del Proceso a Prueba",
                  "mediacion": "Acuerdo de Mediación", "acuerdo": "Acuerdo Conciliatorio"}
TIPO_AUD_LABEL = {
    "audiencia":       "Audiencia contravencional",
    "mediacion":       "Audiencia de mediación",
    "acta_compromiso": "Suscripción acta de compromiso",
    "control_seg":     "Control de seguimiento",
    "reprogramada":    "Audiencia reprogramada",
}
ESTADO_AUD_LABEL = {
    "programada": "Programada", "realizada": "Realizada",
    "ausente": "Ausente", "reprogramada": "Reprogramada", "cancelada": "Cancelada",
}


def _xl_header(writer, sheet, titulo):
    """Fila de título institucional en las primeras tres filas, con merge de columnas."""
    ws = writer.sheets[sheet]
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    # Count how many columns the sheet has (row 4 = data header)
    max_col = max((cell.column for row in ws.iter_rows(min_row=4, max_row=4) for cell in row if cell.value), default=1)

    azul = PatternFill("solid", fgColor="1E2F5E")
    azul2 = PatternFill("solid", fgColor="2E5090")
    font_blanco_big = Font(color="FFFFFF", bold=True, size=12)
    font_blanco_sub = Font(color="FFFFFF", bold=True, size=10)

    ws.insert_rows(1, 3)

    # Row 1: Institution
    ws.cell(1, 1).value = "MINISTERIO PÚBLICO FISCAL DE LA PROVINCIA DE CÓRDOBA"
    ws.cell(1, 1).font = font_blanco_big
    ws.cell(1, 1).fill = azul
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    if max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    # Row 2: Document title
    ws.cell(2, 1).value = titulo
    ws.cell(2, 1).font = font_blanco_sub
    ws.cell(2, 1).fill = azul2
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 17
    if max_col > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    # Row 3: Generation timestamp
    ws.cell(3, 1).value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Sistema SIATC — MPF Córdoba"
    ws.cell(3, 1).font = Font(italic=True, size=8, color="808080")
    ws.cell(3, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 13
    if max_col > 1:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)


def causas_a_excel() -> bytes:
    causas = db.listar_causas(limit=2000)
    if not causas:
        causas = []

    # Hoja 1: Causas
    rows_c = []
    for c in causas:
        rows_c.append({
            "Expediente":      c.get("numero", ""),
            "Imputado/a":      c.get("apellido_nombre", ""),
            "DNI":             c.get("persona_dni", ""),
            "Tipo infracción": TIPOS_INFRACCION.get(c.get("tipo_infraccion",""), {}).get("label", c.get("tipo_infraccion","")),
            "Artículo":        TIPOS_INFRACCION.get(c.get("tipo_infraccion",""), {}).get("articulo", ""),
            "Categoría":       TIPOS_INFRACCION.get(c.get("tipo_infraccion",""), {}).get("categoria", ""),
            "Carril":          CARRIL_LABEL.get(c.get("carril",""), c.get("carril","")),
            "Acción":          c.get("accion", ""),
            "Estado":          c.get("estado", "").capitalize(),
            "Unidad":          c.get("unidad", "").capitalize(),
            "Fiscal":          c.get("fiscal_asignado", ""),
            "Fecha del hecho": c.get("fecha_hecho", "")[:10] if c.get("fecha_hecho") else "",
            "Fecha ingreso":   c.get("created_at", "")[:10] if c.get("created_at") else "",
            "Última actuali.": c.get("updated_at", "")[:10] if c.get("updated_at") else "",
        })
    df_causas = pd.DataFrame(rows_c) if rows_c else pd.DataFrame(
        columns=["Expediente","Imputado/a","DNI","Tipo infracción","Artículo","Categoría",
                 "Carril","Acción","Estado","Unidad","Fiscal",
                 "Fecha del hecho","Fecha ingreso","Última actuali."])

    # Hoja 2: Personas
    personas = db.listar_personas()
    rows_p = []
    for p in personas:
        antec = db.contar_antecedentes(p["id"])
        rows_p.append({
            "DNI":              p.get("dni", ""),
            "Apellido y Nombre":p.get("apellido_nombre", ""),
            "Edad":             p.get("edad", ""),
            "Domicilio":        p.get("domicilio", ""),
            "Teléfono":         p.get("telefono", ""),
            "Antecedentes":     antec,
            "Registrado":       p.get("created_at","")[:10] if p.get("created_at") else "",
        })
    df_personas = pd.DataFrame(rows_p) if rows_p else pd.DataFrame(
        columns=["DNI","Apellido y Nombre","Edad","Domicilio","Teléfono","Antecedentes","Registrado"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_causas.to_excel(writer, sheet_name="Causas", index=False, startrow=3)
        df_personas.to_excel(writer, sheet_name="Personas", index=False, startrow=3)

        # Formato columnas
        for sheet, df in [("Causas", df_causas), ("Personas", df_personas)]:
            ws = writer.sheets[sheet]
            from openpyxl.styles import PatternFill, Font, Alignment
            # Encabezado de columnas (fila 4)
            azul_col = PatternFill("solid", fgColor="2E5090")
            for cell in ws[4]:
                if cell.value:
                    cell.fill = azul_col
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[4].height = 16
            # Ancho automático aproximado
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        # Colores condicionales en Causas — carril column
        ws_c = writer.sheets["Causas"]
        carril_fill = {
            "Mediación":   PatternFill("solid", fgColor="D4EDDA"),
            "Suspensión":  PatternFill("solid", fgColor="FFF3CD"),
            "Proceso pleno": PatternFill("solid", fgColor="F8D7DA"),
        }
        carril_font = {
            "Mediación":   Font(color="155724", bold=True),
            "Suspensión":  Font(color="856404", bold=True),
            "Proceso pleno": Font(color="721C24", bold=True),
        }
        # Find carril column index
        carril_col_idx = None
        estado_col_idx = None
        for cell in ws_c[4]:
            if cell.value == "Carril":
                carril_col_idx = cell.column
            if cell.value == "Estado":
                estado_col_idx = cell.column
        if carril_col_idx:
            for row in ws_c.iter_rows(min_row=5, max_row=ws_c.max_row):
                cv = ws_c.cell(row[0].row, carril_col_idx).value
                if cv in carril_fill:
                    ws_c.cell(row[0].row, carril_col_idx).fill = carril_fill[cv]
                    ws_c.cell(row[0].row, carril_col_idx).font = carril_font[cv]

        _xl_header(writer, "Causas",   "Registro de Causas Contravencionales")
        _xl_header(writer, "Personas", "Registro de Personas Imputadas")

    return buf.getvalue()


def seguimientos_a_excel() -> bytes:
    segs = db.listar_seguimientos()
    today = date.today()

    rows_s = []
    for s in segs:
        condiciones = db.get_condiciones(s["id"])
        prog = db.progress_seguimiento(s["id"])
        try:
            fin = datetime.strptime(s["fecha_fin"], "%Y-%m-%d").date()
            dias_rest = (fin - today).days
        except Exception:
            dias_rest = None

        rows_s.append({
            "ID":              s["id"],
            "Expediente":      s.get("numero", ""),
            "Imputado/a":      s.get("apellido_nombre", ""),
            "DNI":             s.get("dni", ""),
            "Unidad":          s.get("unidad", "").capitalize(),
            "Tipo resolución": TIPO_RES_LABEL.get(s.get("tipo_resolucion",""), s.get("tipo_resolucion","")),
            "Estado":          s.get("estado", "").capitalize(),
            "Fecha inicio":    s.get("fecha_inicio", ""),
            "Fecha fin":       s.get("fecha_fin", ""),
            "Días restantes":  dias_rest,
            "Condiciones total":   prog["total"],
            "Condiciones cumplidas": prog["cumplidas"],
            "% cumplimiento":  prog["pct"],
            "Fiscal":          s.get("fiscal", ""),
        })

    df_seg = pd.DataFrame(rows_s) if rows_s else pd.DataFrame(
        columns=["ID","Expediente","Imputado/a","DNI","Unidad","Tipo resolución",
                 "Estado","Fecha inicio","Fecha fin","Días restantes",
                 "Condiciones total","Condiciones cumplidas","% cumplimiento","Fiscal"])

    # Hoja 2: condiciones detalle
    rows_cd = []
    for s in segs:
        condiciones = db.get_condiciones(s["id"])
        for cond in condiciones:
            acum = db.acumulado_condicion(cond["id"]) if cond["valor_objetivo"] > 0 else 0
            rows_cd.append({
                "Expediente":   s.get("numero", ""),
                "Imputado/a":   s.get("apellido_nombre", ""),
                "Condición":    cond.get("descripcion", ""),
                "Tipo":         cond.get("tipo", "").replace("_", " ").capitalize(),
                "Meta":         cond.get("valor_objetivo", 0) or "",
                "Unidad":       cond.get("unidad", ""),
                "Acumulado":    acum if cond["valor_objetivo"] > 0 else "",
                "Estado":       cond.get("estado", "").capitalize(),
                "Fecha límite": cond.get("fecha_limite", ""),
            })

    df_cond = pd.DataFrame(rows_cd) if rows_cd else pd.DataFrame(
        columns=["Expediente","Imputado/a","Condición","Tipo","Meta","Unidad",
                 "Acumulado","Estado","Fecha límite"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_seg.to_excel(writer,  sheet_name="Seguimientos", index=False, startrow=3)
        df_cond.to_excel(writer, sheet_name="Condiciones",  index=False, startrow=3)

        for sheet, df in [("Seguimientos", df_seg), ("Condiciones", df_cond)]:
            ws = writer.sheets[sheet]
            from openpyxl.styles import PatternFill, Font
            azul_col = PatternFill("solid", fgColor="2E5090")
            for cell in ws[4]:
                if cell.value:
                    cell.fill = azul_col
                    cell.font = Font(color="FFFFFF", bold=True)
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

            # Colores condicionales en col Estado (Seguimientos)
            if sheet == "Seguimientos":
                from openpyxl.styles import PatternFill
                for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
                    estado_cell = None
                    for cell in row:
                        if ws.cell(4, cell.column).value == "Estado":
                            estado_cell = cell
                            break
                    if estado_cell and estado_cell.value:
                        val = str(estado_cell.value).lower()
                        if val == "cumplido":
                            estado_cell.fill = PatternFill("solid", fgColor="C3E6CB")
                        elif val == "incumplido":
                            estado_cell.fill = PatternFill("solid", fgColor="F5C6CB")
                        elif val == "activo":
                            estado_cell.fill = PatternFill("solid", fgColor="FFF3CD")

        _xl_header(writer, "Seguimientos", "Registro de Seguimientos Post-Resolución")
        _xl_header(writer, "Condiciones",  "Detalle de Condiciones por Seguimiento")

    return buf.getvalue()


def audiencias_a_excel() -> bytes:
    """Exporta el listado de audiencias a Excel con dos hojas: futuras y pasadas."""
    from openpyxl.styles import PatternFill, Font
    audiencias = db.listar_audiencias()
    today = date.today().isoformat()

    TIPO_AUD_LABEL_LOCAL = {
        "audiencia":       "Audiencia contravencional",
        "mediacion":       "Audiencia de mediacion",
        "acta_compromiso": "Suscripcion acta de compromiso",
        "control_seg":     "Control de seguimiento",
        "reprogramada":    "Audiencia reprogramada",
    }

    def _row(a):
        return {
            "Fecha":         a.get("fecha", ""),
            "Hora":          a.get("hora", ""),
            "Imputado/a":    a.get("apellido_nombre", ""),
            "DNI":           a.get("dni", ""),
            "Teléfono":      a.get("persona_telefono", ""),
            "Expediente":    a.get("numero", ""),
            "Unidad":        a.get("unidad", "").capitalize(),
            "Carril":        CARRIL_LABEL.get(a.get("carril",""), a.get("carril","")),
            "Tipo":          TIPO_AUD_LABEL_LOCAL.get(a.get("tipo",""), a.get("tipo","")),
            "Lugar":         a.get("lugar",""),
            "Estado":        a.get("estado","").capitalize(),
            "Observaciones": a.get("observaciones",""),
        }

    futuras = [_row(a) for a in audiencias if a.get("fecha","") >= today]
    pasadas = [_row(a) for a in audiencias if a.get("fecha","") <  today]

    cols = ["Fecha","Hora","Imputado/a","DNI","Teléfono","Expediente","Unidad",
            "Carril","Tipo","Lugar","Estado","Observaciones"]

    df_fut = pd.DataFrame(futuras) if futuras else pd.DataFrame(columns=cols)
    df_pas = pd.DataFrame(pasadas) if pasadas else pd.DataFrame(columns=cols)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_fut.to_excel(writer, sheet_name="Proximas",  index=False, startrow=3)
        df_pas.to_excel(writer, sheet_name="Historial", index=False, startrow=3)

        estado_colors = {
            "Programada":   "CCE5FF",
            "Realizada":    "C3E6CB",
            "Ausente":      "F5C6CB",
            "Reprogramada": "FFF3CD",
            "Cancelada":    "E2E3E5",
        }
        azul_col = PatternFill("solid", fgColor="2E5090")

        for sheet, df in [("Proximas", df_fut), ("Historial", df_pas)]:
            ws = writer.sheets[sheet]
            for cell in ws[4]:
                if cell.value:
                    cell.fill = azul_col
                    cell.font = Font(color="FFFFFF", bold=True)
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
                for cell in row:
                    if ws.cell(4, cell.column).value == "Estado" and cell.value:
                        color = estado_colors.get(str(cell.value), None)
                        if color:
                            cell.fill = PatternFill("solid", fgColor=color)

        _xl_header(writer, "Proximas",  "Agenda de Audiencias - Proximas")
        _xl_header(writer, "Historial", "Agenda de Audiencias - Historial")

    return buf.getvalue()
